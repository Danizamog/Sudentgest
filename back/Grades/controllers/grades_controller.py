import os
import json
import httpx
from typing import List, Dict, Optional
from datetime import datetime
from fastapi import UploadFile

# Importar utilidades de supabase (tu archivo existente)
from utils.supabase import (
    get_tenant_from_email,
    get_schema_from_domain,
)

# Variables de entorno (supabase REST)
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


class GradeController:
    """
    Controlador para operaciones con notas.
    Trabaja vía Supabase REST (httpx) y soporta multitenancy por schema.
    """

    @staticmethod
    def _table_name_for_schema(schema: str) -> str:
        # Nombre de la tabla en Supabase por convención: <schema>_notas
        return f"{schema}_notas"

    @staticmethod
    async def list_all_grades(user_email: str) -> List[Dict]:
        """Lista todas las notas del tenant del usuario."""
        domain = get_tenant_from_email(user_email)
        schema = get_schema_from_domain(domain)
        table = GradeController._table_name_for_schema(schema)

        url = f"{SUPABASE_URL}/rest/v1/{table}?select=*"
        headers = {
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        }

        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(url, headers=headers)
            resp.raise_for_status()
            return resp.json()

    @staticmethod
    async def list_grades_by_course(user_email: str, course_id: int) -> List[Dict]:
        """Lista notas de un curso específico (filtra por institution via schema)."""
        domain = get_tenant_from_email(user_email)
        schema = get_schema_from_domain(domain)
        table = GradeController._table_name_for_schema(schema)

        url = f"{SUPABASE_URL}/rest/v1/{table}?course_id=eq.{course_id}&select=*"
        headers = {
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        }

        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(url, headers=headers)
            resp.raise_for_status()
            return resp.json()

    @staticmethod
    async def create_grades_bulk(user_email: str, grades: List[Dict]) -> Dict:
        """
        Inserta múltiples registros de notas.
        grades: lista de dicts con keys: student_id, course_id, grade, (subject opt.)
        """
        domain = get_tenant_from_email(user_email)
        schema = get_schema_from_domain(domain)
        table = GradeController._table_name_for_schema(schema)

        # normalize rows: add timestamp
        now = datetime.utcnow().isoformat()
        to_insert = []
        for g in grades:
            rec = {
                "student_id": g.get("student_id"),
                "course_id": g.get("course_id"),
                "grade": g.get("grade"),
                "subject": g.get("subject") if "subject" in g else None,
                "registered_at": now,
            }
            to_insert.append(rec)

        url = f"{SUPABASE_URL}/rest/v1/{table}"
        headers = {
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        }

        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(url, headers=headers, json=to_insert)
            resp.raise_for_status()
            return {"message": "Notas registradas correctamente", "data": resp.json()}

    @staticmethod
    async def update_grade(user_email: str, grade_id: int, payload: Dict) -> Dict:
        """
        Actualiza una nota por id (sólo dentro del mismo tenant/schema).
        payload: dict con campos a actualizar (por ejemplo: {"grade": 85.5})
        """
        domain = get_tenant_from_email(user_email)
        schema = get_schema_from_domain(domain)
        table = GradeController._table_name_for_schema(schema)

        url = f"{SUPABASE_URL}/rest/v1/{table}?id=eq.{grade_id}"
        headers = {
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        }

        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.patch(url, headers=headers, json=payload)
            resp.raise_for_status()
            return {"message": "Nota actualizada correctamente", "data": resp.json()}

    @staticmethod
    async def delete_grade(user_email: str, grade_id: int) -> Dict:
        """Elimina una nota por id (tenant-aware)."""
        domain = get_tenant_from_email(user_email)
        schema = get_schema_from_domain(domain)
        table = GradeController._table_name_for_schema(schema)

        url = f"{SUPABASE_URL}/rest/v1/{table}?id=eq.{grade_id}"
        headers = {
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        }

        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.delete(url, headers=headers)
            resp.raise_for_status()
            return {"message": "Nota eliminada correctamente"}

    @staticmethod
    async def upload_file_and_create(file: UploadFile, user_email: str) -> Dict:
        """
        Procesa un archivo CSV o XLSX y crea registros en la tabla tenant_notas.
        Formatos esperados (column names, case-insensitive):
          - student_id
          - course_id
          - grade
          - subject  (opcional)
        """
        filename = file.filename.lower()
        # leer bytes del upload
        contents = await file.read()

        rows = []  # cada row será dict con keys: student_id, course_id, grade, subject
        try:
            if filename.endswith(".csv"):
                import io, csv
                text = contents.decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(text))
                for r in reader:
                    # normalizar claves: lower
                    row = {k.strip().lower(): v for k, v in r.items()}
                    rows.append(row)
            elif filename.endswith(".xlsx") or filename.endswith(".xls"):
                from openpyxl import load_workbook
                import io
                wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
                ws = wb.active
                # First row header
                headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row_cells in ws.iter_rows(min_row=2, values_only=True):
                    row = {}
                    for i, h in enumerate(headers):
                        if not h:
                            continue
                        row[h] = row_cells[i]
                    rows.append(row)
            else:
                return {"error": "Formato de archivo no soportado. Use CSV o XLSX."}
        except Exception as e:
            raise Exception(f"Error procesando el archivo: {e}")

        # Convertir a registros normalizados
        normalized = []
        for r in rows:
            try:
                student_id = int(r.get("student_id") or r.get("studentid") or r.get("id"))
                course_id = int(r.get("course_id") or r.get("courseid") or r.get("course"))
                grade_val = r.get("grade") or r.get("nota") or r.get("score")
                grade_num = float(grade_val) if grade_val not in (None, "") else None
                subject = r.get("subject") or r.get("materia")
                normalized.append({
                    "student_id": student_id,
                    "course_id": course_id,
                    "grade": grade_num,
                    "subject": subject,
                    "registered_at": datetime.utcnow().isoformat()
                })
            except Exception:
                # Ignorar filas malformadas
                continue

        # finalmente insertamos en la tabla del tenant
        # necesitamos determinar schema del usuario
        domain = get_tenant_from_email(user_email)
        schema = get_schema_from_domain(domain)
        table = GradeController._table_name_for_schema(schema)

        url = f"{SUPABASE_URL}/rest/v1/{table}"
        headers = {
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        }

        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(url, headers=headers, json=normalized)
            resp.raise_for_status()
            return {"message": "Archivo procesado e importado correctamente", "inserted": resp.json()}
